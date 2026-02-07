import pandas as pd
import re
from typing import List
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

# Debug prints are always enabled - no print override


def convert_gdrive_link(url):
    """
    Convert Google Drive sharing links to direct embeddable image URLs.
    
    Supported formats:
    - https://drive.google.com/file/d/FILE_ID/view?usp=sharing
    - https://drive.google.com/open?id=FILE_ID
    - https://drive.google.com/uc?id=FILE_ID
    - https://drive.usercontent.google.com/download?id=FILE_ID&...
    
    Returns the direct link format that can be used as image src.
    Uses lh3.googleusercontent.com format for reliable browser embedding.
    
    IMPORTANT: The file must be shared with "Anyone with link can view" permissions.
    """
    if not url or not isinstance(url, str):
        return url
    
    url = url.strip()
    
    # Skip if not a Google Drive link
    if 'drive.google.com' not in url and 'drive.usercontent.google.com' not in url:
        print(f"[GDRIVE] Not a Google Drive link: {url[:50]}...")
        return url
    
    file_id = None
    
    # Pattern 1: /file/d/FILE_ID/view
    match = re.search(r'/file/d/([a-zA-Z0-9_-]+)', url)
    if match:
        file_id = match.group(1)
    
    # Pattern 2: ?id=FILE_ID or &id=FILE_ID
    if not file_id:
        match = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', url)
        if match:
            file_id = match.group(1)
    
    # Pattern 3: Already in lh3 format
    if 'lh3.googleusercontent.com' in url:
        print(f"[GDRIVE] Already in lh3 format: {url}")
        return url
    
    if file_id:
        # Use lh3.googleusercontent.com format - this works reliably in browser <img> tags
        # The uc?export=view format often gets blocked by CORS/redirects
        converted = f'https://lh3.googleusercontent.com/d/{file_id}'
        print(f"[GDRIVE] Converted: {url[:50]}... -> {converted}")
        return converted
    
    print(f"[GDRIVE] Could not extract file ID from: {url}")
    # If we couldn't extract file ID, return original
    return url


class ExcelReader:
    """
    This temporary reader will parse excel file containing the roleplay
    In order to extract the needed prompts and queries to be utilised by openai
    In production ideally, this will be taken care by an API
    """
    def __init__(self, path: str, master: dict, image_path: str):
        self.path = path
        self.master = master
        xls = pd.ExcelFile(self.path)
        
        # Find tags and flow sheets - allow any number of sheets, just find the ones we need
        self.tags_sheet = None
        self.flow_sheet = None
        for sheet in xls.sheet_names:
            if "tags" in sheet.lower():
                self.tags_sheet = sheet
            elif "flow" in sheet.lower() and self.flow_sheet is None:
                # Use the first flow sheet found (skip ones with "do not use" in name)
                if "do not use" not in sheet.lower():
                    self.flow_sheet = sheet
        
        # If no valid flow sheet found (all had "do not use"), use the first flow sheet anyway
        if self.flow_sheet is None:
            for sheet in xls.sheet_names:
                if "flow" in sheet.lower():
                    self.flow_sheet = sheet
                    break
        
        if self.tags_sheet == None or self.flow_sheet == None:
            raise ValueError("Cannot find tags or flow sheet in Excel File provided")
        
        self.tag_data = xls.parse(self.tags_sheet)
        self.data = xls.parse(self.flow_sheet, header=None)

        self.image_path = image_path
        image_xls = pd.ExcelFile(self.image_path)
        
        # Find flow sheet in image file - allow any number of sheets
        self.image_flow_sheet = None
        for sheet in image_xls.sheet_names:
            if "flow" in sheet.lower() and self.image_flow_sheet is None:
                # Use the first flow sheet found (skip ones with "do not use" in name)
                if "do not use" not in sheet.lower():
                    self.image_flow_sheet = sheet
        
        # If no valid flow sheet found, use the first flow sheet anyway
        if self.image_flow_sheet is None:
            for sheet in image_xls.sheet_names:
                if "flow" in sheet.lower():
                    self.image_flow_sheet = sheet
                    break
        
        if self.image_flow_sheet == None:
            raise ValueError("Cannot find flow sheet in Excel File provided")
        self.image_data = image_xls.parse(self.image_flow_sheet, header=None)

    def get_all_competencies(self) -> dict:
        """
        Returns all competencies listed in the tags sheet, regardless of 'Enabled'.
        """
        if 'Competency' in self.tag_data.columns and 'Max Score' in self.tag_data.columns:
            mapping_dict = self.tag_data.set_index('Competency')['Max Score'].to_dict()
            return mapping_dict
        return {}
    
    def get_max_scores_from_flow(self) -> dict:
        """
        Calculate the maximum possible score for each competency based on the flow sheet.
        For each interaction, find the BEST (highest) score for each competency,
        then sum those best scores across all interactions.
        
        Returns: dict mapping competency full name -> max possible score
        """
        # Track best scores per competency per interaction
        # Structure: {competency_name: {interaction_num: best_score}}
        competency_interaction_scores = {}
        
        # Find all interactions in the flow sheet
        # Interactions are identified by numeric values in column A (index 0)
        for row_idx in range(len(self.data)):
            interaction_num = self.data.iloc[row_idx, 0]
            
            # Skip non-numeric rows (headers, empty rows, etc.)
            if pd.isna(interaction_num) or not isinstance(interaction_num, (int, float)):
                continue
            
            interaction_num = int(interaction_num)
            
            # The competency row is at row_idx + 1 (one row below the interaction number)
            comp_row_idx = row_idx + 1
            if comp_row_idx >= len(self.data):
                continue
            
            # Read competency cells from columns C, D, E (indices 2, 3, 4) for scores 1, 2, 3
            for col_idx in range(2, 5):  # Columns C, D, E
                if col_idx >= self.data.shape[1]:
                    continue
                    
                cell_value = self.data.iloc[comp_row_idx, col_idx]
                if pd.isna(cell_value) or not cell_value:
                    continue
                
                # Parse competency entries (may have multiple per cell, separated by newlines)
                comp_entries = str(cell_value).split("\n")
                for comp_entry in comp_entries:
                    comp_entry = comp_entry.strip()
                    if not comp_entry:
                        continue
                    
                    # Parse "ABBR LEVEL X:score" format
                    parts = comp_entry.split(":")
                    abbr_key = parts[0].strip()
                    
                    # Get score (after colon, or default to column level)
                    score = col_idx - 1  # Default: col C=1, D=2, E=3
                    if len(parts) > 1:
                        try:
                            score = int(parts[1].strip())
                        except ValueError:
                            pass
                    
                    # Get full competency name from master file
                    if abbr_key in self.master:
                        full_name = self.master[abbr_key].get('name', abbr_key)
                    else:
                        full_name = abbr_key  # Use abbr as fallback
                    
                    # Track the best score for this competency in this interaction
                    if full_name not in competency_interaction_scores:
                        competency_interaction_scores[full_name] = {}
                    
                    current_best = competency_interaction_scores[full_name].get(interaction_num, 0)
                    if score > current_best:
                        competency_interaction_scores[full_name][interaction_num] = score
        
        # Sum the best scores across all interactions for each competency
        max_scores = {}
        for comp_name, interaction_scores in competency_interaction_scores.items():
            max_scores[comp_name] = sum(interaction_scores.values())
        
        return max_scores
    
    def get_system_prompt(self):
        """
        Returns situation description
        """
        return self.data.iloc[0,2]
    
    def get_system_prompt_image(self):
        """Returns the system prompt image URL, converting Google Drive links if needed."""
        image_url = self.image_data.iloc[0,2]
        return convert_gdrive_link(image_url)
    
    def _get_bold_words(self, row: int, col: int) -> List[str]:
        """
        Returns bold words from given cell
        Supports both .xls (xlrd) and .xlsx (openpyxl) formats
        """
        try:
            # Check if file is .xls or .xlsx
            if self.path.endswith('.xls'):
                # Use xlrd for .xls files
                if not XLRD_AVAILABLE:
                    print(f"Warning: xlrd not available for .xls file: {self.path}")
                    return []
                    
                workbook = xlrd.open_workbook(self.path, formatting_info=True)
                sheet = workbook.sheet_by_name(self.flow_sheet)
                cell_value = sheet.cell_value(row, col)
                rich_text_runlist = sheet.rich_text_runlist_map.get((row, col))
                
                if rich_text_runlist is None:
                    return []
                    
                bold_phrases = []
                if rich_text_runlist:
                    rich_text_runlist.append((len(cell_value), None))
                    for i in range(len(rich_text_runlist)-1):
                        font = workbook.font_list[rich_text_runlist[i][1]]
                        if font.bold:
                            bold_phrases.append(cell_value[rich_text_runlist[i][0]: rich_text_runlist[i+1][0]].strip())
                return bold_phrases
                
            elif self.path.endswith('.xlsx'):
                # Use openpyxl for .xlsx files
                if not OPENPYXL_AVAILABLE:
                    print(f"Warning: openpyxl not available for .xlsx file: {self.path}")
                    return []
                    
                workbook = load_workbook(self.path, data_only=False)
                sheet = workbook[self.flow_sheet]
                
                # openpyxl uses 1-based indexing
                cell = sheet.cell(row=row+1, column=col+1)
                cell_value = str(cell.value) if cell.value else ""
                
                bold_phrases = []
                
                # Check if cell has rich text formatting
                if hasattr(cell.value, '__iter__') and not isinstance(cell.value, str):
                    # Rich text is a list of tuples in openpyxl
                    for text_obj in cell.value:
                        if hasattr(text_obj, 'font') and text_obj.font and text_obj.font.b:
                            bold_phrases.append(text_obj.text.strip())
                elif cell.font and cell.font.b:
                    # Entire cell is bold
                    bold_phrases.append(cell_value.strip())
                
                workbook.close()
                return bold_phrases
            else:
                print(f"Warning: Unsupported file format: {self.path}")
                return []
                
        except Exception as e:
            # Don't print warning for .xls files - it's expected
            if not self.path.endswith('.xls'):
                print(f"Warning: Could not extract bold formatting from cell ({row}, {col}): {e}")
            return []
    
    def get_interaction(self, current_interaction_number: int) -> dict:
        """
        Given the interaction number, returns the current interaction responses
        Both the player and computer responses in order to create an effective prompt
        """

        matching_rows = self.data.loc[self.data[0] == current_interaction_number]
        if len(matching_rows.index) == 0:
            print(f"❌ ERROR: No rows found with interaction number {current_interaction_number}")
            return False
        
        if len(matching_rows.index) > 1:
            print(f"⚠️ WARNING: Multiple rows found with interaction number {current_interaction_number}!")
            print(f"   Found at DataFrame indices: {list(matching_rows.index)}")
            print(f"   (Excel rows: {[idx + 1 for idx in matching_rows.index]})")
            print(f"   Using first match: index {matching_rows.index[0]} (Excel row {matching_rows.index[0] + 1})")
        
        current_index = matching_rows.index[0]
        
        print(f"\n📖 READING INTERACTION #{current_interaction_number}")
        print(f"   Excel Row for interaction: {current_index + 1}")  # +1 for Excel row numbering
        print(f"   Player choices row: {current_index + 1}")
        print(f"   Competency row: {current_index + 2}")
        print(f"   Computer response row: {current_index + 3}")
        print(f"   DataFrame columns count: {self.data.shape[1]}")
        
        # Try to get tip from column 5 (F), if it exists (column F is index 5, so need at least 6 columns)
        tip = None
        try:
            if self.data.shape[1] > 5:  # Need MORE than 5 columns to access index 5
                tip = self.data.iloc[current_index, 5]
                if pd.isna(tip):
                    tip = None
                else:
                    print(f"   Tip found: {tip}")
            else:
                print(f"   No tip column available (only {self.data.shape[1]} columns)")
        except (IndexError, KeyError) as e:
            print(f"   Could not read tip column: {e}")
            tip = None
        
        player = self.data.iloc[current_index, 2:].tolist()[:3]
        keywords = [self._get_bold_words(current_index, x) for x in range(2,5)]
        competency = self._process_choice_competencies(self.data.iloc[current_index+1, 2:].tolist()[:3], self.master)
        comp = self.data.iloc[current_index+2, 2:].tolist()[:3]
        
        print(f"   Player choices: {player}")
        print(f"   Computer responses RAW: {comp}")
        print(f"   Computer responses TYPES: {[type(x) for x in comp]}")
        
        # Convert any non-string values to strings (handles numbers, NaN, etc.)
        comp_cleaned = []
        for idx, item in enumerate(comp):
            if pd.isna(item):
                print(f"   ⚠️ WARNING: Computer response {idx+1} (score {idx+1}) is empty/NaN!")
                comp_cleaned.append("")
            elif isinstance(item, (int, float)):
                print(f"   ⚠️ WARNING: Computer response {idx+1} (score {idx+1}) is a number: {item}")
                comp_cleaned.append(str(item))
            else:
                comp_cleaned.append(str(item) if item else "")
        
        comp = comp_cleaned
        print(f"   Computer responses CLEANED: {[x[:50] if x else 'EMPTY' for x in comp]}")  # First 50 chars
        
        # Extract character names and gender from computer response
        # Two formats supported:
        # 1. Team roleplay: "Bheem (M): Yes Sir | Satyam (M): All okay" (multi-speaker)
        # 2. Single roleplay: Column B = "other (M)" or "other (F)" (single speaker)
        characters = []
        character = None
        gender_marker = None  # For single-speaker roleplays
        
        # Check Column B (index 1) for single-speaker gender marker: "other (M)" or "other (F)"
        column_b_value = self.data.iloc[current_index+2, 1]  # Row 3 (computer response), Column B
        if pd.notna(column_b_value):
            column_b_str = str(column_b_value).strip().lower()
            print(f"🔍 EXCEL DEBUG: Row {current_index+3} Column B value = '{column_b_value}' (lowercased: '{column_b_str}')")
            import re
            # Look for (M), (F), (Male), (Female) in Column B
            if '(m)' in column_b_str or '(male)' in column_b_str:
                gender_marker = 'male'
                print(f"✅ GENDER MARKER DETECTED: MALE from Column B")
            elif '(f)' in column_b_str or '(female)' in column_b_str:
                gender_marker = 'female'
                print(f"✅ GENDER MARKER DETECTED: FEMALE from Column B")
            else:
                print(f"⚠️ No gender marker found in Column B (expected '(M)' or '(F)')")
        
        # Extract character names from dialogue text for team roleplays
        for response_text in comp:
            if response_text and not pd.isna(response_text):
                text = str(response_text)
                # Find all "Name:" patterns
                matches = re.findall(r'([A-Z][a-zA-Z]+):', text)
                for name in matches:
                    if name and name not in characters:
                        characters.append(name)
        
        # Set primary character as first one found
        if characters:
            character = characters[0]
        
        return {
            "player": player, 
            "comp": comp, 
            "competencies": competency, 
            "tip": tip, 
            "keywords": keywords, 
            "character": character, 
            "characters": characters,
            "gender_marker": gender_marker  # For single-speaker roleplays
        }
    
    def get_images(self, current_interaction_number: int):
        """
        Get images for the current interaction.
        Converts Google Drive links to direct viewable URLs.
        """
        placeholder = "https://developers.elementor.com/docs/assets/img/elementor-placeholder-image.png"
        try:
            current_index = self.image_data.loc[self.image_data[0] == current_interaction_number]
            if len(current_index.index) == 0:
                return {"images": [placeholder, placeholder, placeholder]}
            current_index = current_index.index[0]
            images = self.image_data.iloc[current_index+2, 2:].tolist()[:3]
            
            # Convert Google Drive links to direct URLs
            converted_images = [convert_gdrive_link(img) if isinstance(img, str) else placeholder for img in images]
            
            # Ensure we always have 3 images
            while len(converted_images) < 3:
                converted_images.append(placeholder)
            
            return {"images": converted_images}
        except Exception as e:
            print(f"Error getting images for interaction {current_interaction_number}: {e}")
            return {"images": [placeholder, placeholder, placeholder]}
        
    def _process_choice_competencies(self, options: List[str], descriptions: dict) -> List[dict]:
        """
        Input: List of competency strings (e.g., "MOTVN LEVEL 2:2"), master dict of competencies
        Output: RETURNS COMPETENCIES FOR THIS INTERACTION AS LIST OF DICTS
                Each dict contains: name, description, examples, expected_score, column_level
                
        NOTE: Returns ALL competencies from ALL columns (1, 2, 3) so that we can
        pick the correct scores based on which response the user matched.
        """
        print(f"\n📋 PROCESSING COMPETENCIES FROM EXCEL:")
        print(f"   Raw competency data from columns C/D/E:")
        for i, opt in enumerate(options):
            print(f"   Column {i+1} (score {i+1}): {opt}")
        
        final_list = []
        
        for col_idx, c_string in enumerate(options):
            # col_idx: 0=score1, 1=score2, 2=score3
            column_level = col_idx + 1  # 1, 2, or 3
            
            if pd.isna(c_string) or not c_string:
                continue
                
            comps = str(c_string).split("\n")
            for comp in comps:
                comp = comp.strip()
                if not comp:
                    continue
                    
                # Parse "ABBR LEVEL X:score" format
                # e.g., "MOTVN LEVEL 2:2" -> key="MOTVN LEVEL 2", score=2
                parts = comp.split(":")
                key = parts[0].strip()
                
                # Extract the score if provided (after the colon)
                comp_score = column_level  # Default to column level
                if len(parts) > 1:
                    try:
                        comp_score = int(parts[1].strip())
                    except ValueError:
                        comp_score = column_level
                
                if key not in descriptions:
                    # Better error message showing what's wrong
                    available_keys = list(descriptions.keys())
                    raise ValueError(f"Could not find competency '{key}' in master file. Check spelling and spacing. Available: {available_keys}")
                
                comp_data = descriptions[key].copy()  # Get name, description, examples
                comp_data["expected_score"] = comp_score  # Score value for this competency
                comp_data["column_level"] = column_level  # Which response column (1, 2, or 3)
                
                # Add ALL competencies (including same name from different columns)
                # The openai.py will filter based on matched column
                final_list.append(comp_data)
                    
        return final_list

    def get_next_interaction(self, interaction_number: int, score: int):
        """
        Considers the input and returns the accurate next interaction 
        (including skipping interactions as mentioned in the excel sheet)
        Returns -1 if interaction is over (changed from False)
        """
        current_index = self.data.loc[self.data[0] == interaction_number].index[0]
        
        if score <= 0 or score > 3:
            raise ValueError("Invalid Score")
        
        print(f"\n🔍 GET_NEXT_INTERACTION: Current interaction={interaction_number}, Score={score}")
        print(f"   Current DataFrame index: {current_index} (Excel row {current_index + 1})")
        
        # Debug: Show the structure of this interaction block
        print(f"   📋 INTERACTION BLOCK STRUCTURE:")
        print(f"      Row {current_index} (Excel {current_index+1}): Player row - Col A = {self.data.iloc[current_index, 0]}")
        print(f"      Row {current_index+1} (Excel {current_index+2}): Competency row - Col A = {self.data.iloc[current_index+1, 0] if current_index+1 < len(self.data) else 'OUT OF BOUNDS'}")
        print(f"      Row {current_index+2} (Excel {current_index+3}): Computer row - Col A = {self.data.iloc[current_index+2, 0] if current_index+2 < len(self.data) else 'OUT OF BOUNDS'}")
        if current_index+3 < len(self.data):
            print(f"      Row {current_index+3} (Excel {current_index+4}): Action row - Col A = {self.data.iloc[current_index+3, 0]}")
            print(f"         Action row full contents: {self.data.iloc[current_index+3, :5].tolist()}")
        else:
            print(f"      Row {current_index+3} (Excel {current_index+4}): OUT OF BOUNDS - End of data")
            return -1
        
        # Check the next row (current_index + 3) to see if it has an interaction number
        next_row_value = self.data.iloc[current_index+3, 0]
        print(f"   Checking for next interaction number in row {current_index+3}: '{next_row_value}'")
        
        if pd.isnull(next_row_value):
            # No interaction number in next row, check the action cell
            action_cell_col = 1 + score  # Column B=2, C=3, D=4 for scores 1, 2, 3
            todo = self.data.iloc[current_index+3, action_cell_col]
            todo_original = str(todo)
            print(f"   Action cell (row {current_index+3}, col {action_cell_col}): '{todo_original}'")
            
            # Handle NaN/None values - the 4th row might be a blank separator row
            # In that case, look for the next interaction in the 5th row (current_index + 4)
            if pd.isna(todo):
                print(f"   ⚠️ Action cell is NaN/None - checking if this is a blank separator row...")
                # Check all cells in row current_index+3 to see if it's completely blank
                row_data = self.data.iloc[current_index+3, :]
                is_blank_row = all(pd.isna(cell) or str(cell).strip() == '' for cell in row_data)
                
                if is_blank_row:
                    # This is a blank separator row, look for next interaction in row current_index+4
                    print(f"   ℹ️ Row {current_index+3} is a blank separator row")
                    try:
                        next_int_value = self.data.iloc[current_index+4, 0]
                        if pd.notna(next_int_value):
                            next_int = int(next_int_value)
                            print(f"   ➡️ Found next interaction in row {current_index+4}: {next_int}")
                            return next_int
                        else:
                            print(f"   ❌ Row {current_index+4} column A is also empty - ENDING roleplay")
                            return -1
                    except (IndexError, ValueError) as e:
                        print(f"   ❌ Error finding next interaction: {e} - ENDING roleplay")
                        return -1
                else:
                    # Row is not completely blank but action cell for this score is empty
                    # Try to find next interaction
                    print(f"   ⚠️ Action cell for score {score} is empty but row has other data")
                    try:
                        next_int_value = self.data.iloc[current_index+4, 0]
                        if pd.notna(next_int_value):
                            next_int = int(next_int_value)
                            print(f"   ➡️ Found next interaction in row {current_index+4}: {next_int}")
                            return next_int
                        else:
                            print(f"   ❌ No next interaction found - ENDING roleplay")
                            return -1
                    except (IndexError, ValueError) as e:
                        print(f"   ❌ Error finding next interaction: {e} - ENDING roleplay")
                        return -1
            
            todo_processed = str(todo).lower().strip()
            todo_processed = re.sub('[^A-Za-z0-9]+', ' ', todo_processed).strip()
            print(f"   Processed action text: '{todo_processed}'")
            
            # Check for goto instruction (e.g., "Go to row 24", "goto row 5", "Go to 24")
            if any(str.isdigit(c) for c in todo_processed):
                extract = todo_processed.split(" ")
                print(f"   Found digits in action cell, split words: {extract}")
                goto_row_number = None
                for e in extract:
                    if e.isnumeric():
                        goto_row_number = int(e)
                        print(f"   ✅ Found goto ROW number: {goto_row_number}")
                        break
                if goto_row_number == None:
                    print(f"   ❌ Could not extract goto number - ENDING roleplay (returning -1)")
                    return -1
                
                # IMPORTANT: goto_row_number is an EXCEL ROW, not an interaction number
                # We need to find which interaction number is at that Excel row
                # Excel rows are 1-indexed, pandas DataFrame is 0-indexed
                # So Excel row 9 = DataFrame index 8
                try:
                    goto_excel_index = goto_row_number - 1  # Convert to 0-indexed
                    print(f"   Converting Excel row {goto_row_number} to DataFrame index {goto_excel_index}")
                    
                    # Get the interaction number at that row (column A = index 0)
                    goto_interaction_number = self.data.iloc[goto_excel_index, 0]
                    
                    if pd.isna(goto_interaction_number):
                        print(f"   ❌ No interaction number at Excel row {goto_row_number} - ENDING roleplay (returning -1)")
                        return -1
                    
                    goto_interaction_number = int(goto_interaction_number)
                    print(f"   ➡️ Excel row {goto_row_number} = INTERACTION {goto_interaction_number}")
                    print(f"   ➡️ GOTO interaction {goto_interaction_number}")
                    return goto_interaction_number
                except Exception as e:
                    print(f"   ❌ Error finding interaction at Excel row {goto_row_number}: {e}")
                    print(f"   Treating {goto_row_number} as interaction number (fallback)")
                    return goto_row_number
            elif "end" in todo_processed:
                # "End Scenario" found - but check if there's actually a next interaction
                # In some Excel formats, "End Scenario" is just a label and the next interaction follows
                print(f"   ⚠️ Found 'end' keyword in action cell for score {score}")
                print(f"   📋 Checking all action cells in this row:")
                for col in range(1, 5):  # Columns B, C, D, E
                    cell_val = self.data.iloc[current_index+3, col] if current_index+3 < len(self.data) else 'OUT OF BOUNDS'
                    print(f"      Column {col}: '{cell_val}'")
                
                # Check if the next row has a new interaction - if so, go there instead of ending
                try:
                    if current_index+4 < len(self.data):
                        next_row_int = self.data.iloc[current_index+4, 0]
                        print(f"   📋 Checking row after action row: index {current_index+4} (Excel row {current_index+5}), Col A = '{next_row_int}'")
                        if pd.notna(next_row_int) and isinstance(next_row_int, (int, float)):
                            next_int = int(next_row_int)
                            print(f"   ✅ Found next interaction #{next_int} after 'End Scenario' - continuing to it!")
                            return next_int
                except Exception as e:
                    print(f"   ⚠️ Error checking for next interaction: {e}")
                
                # No next interaction found - truly end the roleplay
                print(f"   ❌ No next interaction found - ENDING roleplay (returning -1)")
                return -1
            else:
                # No clear instruction, try to go to next numbered interaction
                try:
                    next_int = int(self.data.iloc[current_index+4, 0])
                    print(f"   ➡️ No clear action, going to next numbered row: {next_int}")
                    return next_int
                except:
                    print(f"   ❌ Cannot find next interaction - ENDING roleplay (returning -1)")
                    return -1
        else:
            # Next row has an interaction number, go there
            next_int = int(next_row_value)
            print(f"   ➡️ Next row has interaction number, going to: {next_int}")
            return next_int